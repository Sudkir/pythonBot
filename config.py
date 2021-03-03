# Библиотека для записи и чтения Excel файлов
import openpyxl
# Принимает имя файла в качестве аргумента и возвращает объект рабочей книги
from openpyxl import load_workbook
# Конвертирует индексы столбцов (3 -> 'C')
from openpyxl.utils import get_column_letter
# Конвертирует индексы столбцов ('A' -> 1)
from openpyxl.utils import column_index_from_string


path_to_file = 'merged_tmp.xlsx'  # путь к файлу
# открываю книгу
wb = load_workbook(path_to_file)
# получаю номера листов
wb.sheetnames
# создаю переменную для 1го листа тк именно там находится расписание
sheet = wb['Лист1']


def fullSchedule(group_number, day_week):

    day_rows = [13, 33, 53, 73, 93, 113]
    shedule = ""
    if day_week == "понедельник" or day_week == "пн":
        shedule = print_schedule(day_rows[0], find_group(group_number, path_to_file))
    elif day_week == "вторник" or day_week == "вт":
        shedule = print_schedule(day_rows[1], find_group(group_number, path_to_file))
    elif day_week == "среда" or day_week == "ср":
        shedule = print_schedule(day_rows[2], find_group(group_number, path_to_file))
    elif day_week == "четверг" or day_week == "чт":
        shedule = print_schedule(day_rows[3], find_group(group_number, path_to_file))
    elif day_week == "пятница" or day_week == "пт":
        shedule = print_schedule(day_rows[4], find_group(group_number, path_to_file))
    elif day_week == "суббота" or day_week == "сб":
        shedule = print_schedule(day_rows[5], find_group(group_number, path_to_file))
    elif day_week == "воскресенье" or day_week == "вс":
        shedule = "Выходной"
    else:
        shedule = "Повторите дни недели:)"
    return shedule


def find_group(search_text, patch_to_file):
    wb = openpyxl.load_workbook(path_to_file)  # Грузим наш прайс-лист
    sheets_list = wb.sheetnames  # Получаем список всех листов в файле
    sheet_active = wb[sheets_list[0]]  # Начинаем работать с самым первым
    column_max = sheet_active.max_column  # Получаем количество столбцов

    row_min = 12  # Переменная, отвечающая за номер строки
    column_min = 1  # Переменная, отвечающая за номер столбца
    while column_min <= column_max:

        word_column = get_column_letter(column_min)  # Получаем численный номер столбца

        word_cell = str(word_column) + str(row_min)  # Склеиваем номер столбца и строки

        data_from_cell = sheet_active[word_cell].value  # Получаем значение ячейки с заданными координатами
        data_from_cell = str(data_from_cell)  # Получаем строковое предстваление ячейки с заданными координатами
        # Сравниваем с искомым текстом
        result = False
        if search_text.lower() == data_from_cell.lower():
            result = True
        # Если искомое значение совпадает со значением ячейки, то возвращаем координаты ячейки
        if result:
            column_rows = [column_index_from_string(word_column), row_min]
            return column_rows
        column_min = column_min + 1
    return [None, None]

def print_schedule(rows, koord_groop):
    if koord_groop[0] == None:
        return "Нет такой группы в нашей базе"
    result = ""
    # Запустим цикл по количеству пар в день
    for k in range(5):
        flag = False
        str_lesson = "" # В эту перменную запишем все пары дня
        rows_l = rows + 4 # После каждой итерации прибаляем количество строк одной пары
        for i in range(rows, rows_l):
            lesson = sheet.cell(row=i, column=koord_groop[0]).value
            if lesson != None: # Если строка пары не пустая, то добавляем ее
                flag = True
                str_lesson += lesson + "\n"
        if flag:
            time = sheet.cell(row=rows, column=2).value
            if time != None:
                result += time + "\n" + str_lesson
            else:
                result += str_lesson
        rows += 4
    return result




